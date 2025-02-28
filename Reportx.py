import os
import pandas as pd
import glob
import re
from openpyxl.styles import Font, PatternFill

# Define the folder path
folder_path = r"C:\py\Program\Data"

# Function to normalize column names for better matching
def normalize_column_name(name):
    return re.sub(r'\s+', '', name).lower()

# Function to find the best match for a column name
def find_column_match(df, column_name):
    # Normalize the target column name
    norm_target = normalize_column_name(column_name)
    
    # Create a dictionary of normalized column names to actual column names
    norm_to_actual = {normalize_column_name(col): col for col in df.columns}
    
    # Check for exact match after normalization
    if norm_target in norm_to_actual:
        return norm_to_actual[norm_target]
    
    # Check for partial matches
    possible_matches = [actual for norm, actual in norm_to_actual.items() 
                       if norm_target in norm or norm in norm_target]
    
    if possible_matches:
        return possible_matches[0]
    
    return None

# Function to read a file based on its extension
def read_file(file_path):
    _, ext = os.path.splitext(file_path)
    
    if ext.lower() in ['.csv']:
        try:
            return pd.read_csv(file_path)
        except UnicodeDecodeError:
            try:
                return pd.read_csv(file_path, encoding='latin1')
            except Exception as e:
                print(f"Error reading {file_path}: {e}")
                return pd.DataFrame()
    elif ext.lower() in ['.xlsx', '.xls']:
        try:
            return pd.read_excel(file_path)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return pd.DataFrame()
    elif ext.lower() in ['.txt', '.dat']:
        # Try with different separators
        for sep in ['\t', ',', ';', '|']:
            try:
                df = pd.read_csv(file_path, sep=sep)
                if len(df.columns) > 1:  # If more than one column, separator worked
                    return df
            except:
                continue
        print(f"Could not determine separator for {file_path}")
        return pd.DataFrame()
    else:
        print(f"Unsupported file format: {file_path}")
        return pd.DataFrame()

# Function to find AE and PT files
def find_files(folder, file_type):
    all_files = []
    for extension in ['*.csv', '*.xlsx', '*.xls', '*.txt', '*.dat']:
        search_pattern = os.path.join(folder, f"*{file_type}*{extension}")
        all_files.extend(glob.glob(search_pattern))
    return all_files

# Find AE and PT files
ae_files = find_files(folder_path, "AE")
pt_files = find_files(folder_path, "PT")

print(f"Found {len(ae_files)} AE files: {ae_files}")
print(f"Found {len(pt_files)} PT files: {pt_files}")

# Read all AE files and concatenate them
ae_data_frames = []
for file in ae_files:
    print(f"Reading AE file: {file}")
    df = read_file(file)
    if not df.empty:
        print(f"  - Read {len(df)} rows with {len(df.columns)} columns")
        ae_data_frames.append(df)

# Concatenate all AE data frames
if ae_data_frames:
    ae_data = pd.concat(ae_data_frames, ignore_index=True)
    print(f"Total AE records: {len(ae_data)}")
else:
    ae_data = pd.DataFrame()
    print("No AE data loaded.")

# Read all PT files and concatenate them
pt_data_frames = []
for file in pt_files:
    print(f"Reading PT file: {file}")
    df = read_file(file)
    if not df.empty:
        print(f"  - Read {len(df)} rows with {len(df.columns)} columns")
        pt_data_frames.append(df)

# Concatenate all PT data frames
if pt_data_frames:
    pt_data = pd.concat(pt_data_frames, ignore_index=True)
    print(f"Total PT records: {len(pt_data)}")
else:
    pt_data = pd.DataFrame()
    print("No PT data loaded.")

# Extract required fields from AE data
if not ae_data.empty:
    try:
        # Define required columns and look for matching columns
        ae_required_cols = {
            'Activity Seq': None,
            'Project': None,
            'Project Description': None,
            'Activity': None,
            'Activity Description': None,
            'Estimated Revenue': None,
            'Estimated Cost': None
        }
        
        # Find matching columns for each required column
        for req_col in ae_required_cols:
            match = find_column_match(ae_data, req_col)
            if match:
                ae_required_cols[req_col] = match
                print(f"Matched '{req_col}' to '{match}' in AE data")
            else:
                print(f"Warning: No match found for '{req_col}' in AE data")
        
        # Check if we found Activity Seq (primary key)
        if ae_required_cols['Activity Seq'] is None:
            print("Error: Could not find 'Activity Seq' column in AE data")
            ae_extract = pd.DataFrame()
        else:
            # Create a new DataFrame with matched columns
            ae_extract = pd.DataFrame()
            
            for req_col, actual_col in ae_required_cols.items():
                if actual_col:
                    ae_extract[req_col] = ae_data[actual_col]
                else:
                    # Create empty column if no match was found
                    ae_extract[req_col] = pd.NA
            
            # Group by Activity Seq to handle potential duplicates
            ae_extract = ae_extract.groupby('Activity Seq').first().reset_index()
            print(f"Extracted {len(ae_extract)} unique Activity Seq records from AE data")
    except Exception as e:
        print(f"Error processing AE data: {e}")
        ae_extract = pd.DataFrame()
else:
    print("No AE data found.")
    ae_extract = pd.DataFrame()

# Aggregate actual costs from PT data by Activity Seq
if not pt_data.empty:
    # Look for Activity Seq in PT data
    activity_seq_col = find_column_match(pt_data, 'Activity Seq')
    
    if not activity_seq_col:
        print("Error: Could not find 'Activity Seq' column in PT data")
        pt_grouped = pd.DataFrame()
    else:
        print(f"Using '{activity_seq_col}' as the Activity Seq column for PT data")
        
        # Try different possible column names for actual costs
        cost_columns = ['Total Internal Price', 'Internal Price', 'Sales Amount', 
                        'Sales Price', 'Internal Amount']
        
        cost_column = None
        for col in cost_columns:
            match = find_column_match(pt_data, col)
            if match:
                cost_column = match
                break
        
        if cost_column:
            print(f"Using '{cost_column}' as the cost column for PT data")
            
            # Aggregating by Activity Seq
            pt_grouped = pt_data.groupby(activity_seq_col).agg({
                cost_column: 'sum'
            }).reset_index()
            
            # Rename columns for consistency
            pt_grouped.rename(columns={activity_seq_col: 'Activity Seq', 
                                      cost_column: 'Actual Cost'}, inplace=True)
            
            print(f"Aggregated actual costs for {len(pt_grouped)} Activity Seq records from PT data")
        else:
            print(f"No suitable cost column found in PT data")
            pt_grouped = pd.DataFrame()
else:
    print("No PT data found.")
    pt_grouped = pd.DataFrame()

# Merge AE and PT data on Activity Seq
if not ae_extract.empty and not pt_grouped.empty:
    # Left join to keep all Activity Seq from AE data
    final_report = pd.merge(ae_extract, pt_grouped, on='Activity Seq', how='left')
    
    # Fill NaN values in Actual Cost with 0
    final_report['Actual Cost'] = final_report['Actual Cost'].fillna(0)
    
    # Calculate variance
    final_report['Cost Variance'] = final_report['Estimated Cost'] - final_report['Actual Cost']
    
    print(f"Created final report with {len(final_report)} records")
elif not ae_extract.empty:
    # If only AE data is available
    print("Only AE data available for the report")
    final_report = ae_extract
    final_report['Actual Cost'] = 0
    final_report['Cost Variance'] = final_report['Estimated Cost']
elif not pt_grouped.empty:
    # If only PT data is available
    print("Only PT data available for the report")
    final_report = pt_grouped
    # Add missing columns
    for col in ['Project', 'Project Description', 'Activity', 'Activity Description', 
               'Estimated Revenue', 'Estimated Cost']:
        final_report[col] = pd.NA
    final_report['Cost Variance'] = pd.NA
else:
    print("No data available for the report.")
    final_report = pd.DataFrame()

# Write the final report to Excel if data is available
if not final_report.empty:
    output_file = os.path.join(folder_path, "reportX.xlsx")
    
    # Sort the report by Activity Seq
    if 'Activity Seq' in final_report.columns:
        final_report.sort_values('Activity Seq', inplace=True)
    
    # Create a Pandas Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_report.to_excel(writer, sheet_name='Activity Report', index=False)
        
        # Apply formatting
        try:
            workbook = writer.book
            worksheet = writer.sheets['Activity Report']
            
            # Style the header row
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Auto-adjust columns' width
            for i, col in enumerate(final_report.columns):
                col_width = max(
                    final_report[col].astype(str).map(len).max(),
                    len(str(col))
                ) + 2
                column_letter = chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26)
                worksheet.column_dimensions[column_letter].width = col_width
            
            # Format numeric columns
            for row in range(2, len(final_report) + 2):
                for i, col in enumerate(final_report.columns):
                    column_letter = chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26)
                    cell = worksheet[f"{column_letter}{row}"]
                    
                    # Apply currency format to cost/revenue columns
                    if col in ['Estimated Cost', 'Estimated Revenue', 'Actual Cost', 'Cost Variance']:
                        cell.number_format = '$#,##0.00'
        except Exception as e:
            print(f"Warning: Could not apply formatting to Excel file: {e}")
    
    print(f"Report successfully created: {output_file}")
else:
    print("No data to write to the report.")

print("Process completed.")

import pandas as pd
import os
import sys
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time

print("Script execution started")

def create_project_report():
    try:
        print("\n===== PROJECT REPORT GENERATION STARTED =====")
        
        # Define file paths
        data_dir = r'C:\py\Program\Data'
        ae_file = os.path.join(data_dir, 'ae.xlsx')
        p_file = os.path.join(data_dir, 'P.xlsx')
        pt_file = os.path.join(data_dir, 'PT.xlsx')
        
        # Verify files exist
        print(f"Checking for required files...")
        for file_path, file_name in [(ae_file, "AE"), (p_file, "P"), (pt_file, "PT")]:
            if os.path.exists(file_path):
                print(f"  ✓ {file_name} file found: {file_path}")
            else:
                print(f"  ✗ {file_name} file NOT found: {file_path}")