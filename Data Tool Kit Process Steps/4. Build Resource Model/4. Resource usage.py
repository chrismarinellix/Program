import os
from openpyxl import load_workbook
import time

def update_excel_file():
    # Define the file path
    file_path = r"C:\Users\chris.marinelli\OneDrive - Vysus Group\Documents - Energy - Power Engineering\Project Management\Haughton Proejct GRD87.xlsm"
    
    # Check if the file exists
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        return
    
    # First attempt with openpyxl
    try:
        # Load the workbook with keep_vba=True to preserve macros
        workbook = load_workbook(file_path, keep_vba=True)
        
        # Check if 'Schedule' sheet exists
        if 'Schedule' not in workbook.sheetnames:
            print("Error: 'Schedule' sheet not found in the workbook")
            return
        
        # Access the Schedule sheet
        schedule_sheet = workbook['Schedule']
        
        # Set value in cell B14 to 'X'
        schedule_sheet['B14'] = 'X'
        
        # Save the workbook
        workbook.save(file_path)
        
        print(f"Successfully updated cell B14 in the Schedule tab of {file_path}")
        return  # Success, exit the function
        
    except PermissionError:
        print("File is open in Excel. Attempting alternative method...")
    except Exception as e:
        print(f"An error occurred with openpyxl: {str(e)}")
    
    # Second attempt with xlwings if available
    try:
        import xlwings as xw
        
        print("Trying with xlwings...")
        
        # Try to connect to an existing Excel instance
        try:
            app = xw.apps.active
        except:
            app = xw.App(visible=False)
        
        try:
            # Open the workbook in the existing Excel instance
            wb = app.books.open(file_path)
            sheet = wb.sheets['Schedule']
            
            # Set the value
            sheet.range('B14').value = 'X'
            
            # Save without closing
            wb.save()
            print(f"Successfully updated cell B14 using xlwings")
            
        finally:
            # Don't quit Excel if it was already running
            if app.pid is not None:
                app.books[file_path].close()
                if len(app.books) == 0:
                    app.quit()
        
        return  # Success, exit the function
        
    except ImportError:
        print("xlwings not available. Installing it might help with editing open files.")
    except Exception as e:
        print(f"Error with xlwings approach: {str(e)}")
    
    # Third attempt: Save to a temporary file as a last resort
    try:
        # Generate a temporary filename
        temp_path = file_path.replace('.xlsm', '_temp.xlsm')
        
        # Load and modify
        workbook = load_workbook(file_path, keep_vba=True)
        schedule_sheet = workbook['Schedule']
        schedule_sheet['B14'] = 'X'
        
        # Save to temp file
        workbook.save(temp_path)
        
        print(f"File is locked. Changes saved to temporary file: {temp_path}")
        print("Please manually replace the original file with the temporary file when possible.")
        
    except Exception as e:
        print(f"All attempts failed. Final error: {str(e)}")

if __name__ == "__main__":
    update_excel_file()