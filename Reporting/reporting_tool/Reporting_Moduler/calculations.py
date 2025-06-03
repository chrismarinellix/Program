import sys
import os
import re # Import re for sheet name sanitization

# START OF IMPORT FIX
# Get the absolute path of the directory where this script (calculations.py) is located
script_dir = os.path.dirname(os.path.abspath(__file__))

# Add this directory to the Python path (at the beginning)
# This ensures that Python will look in this script's directory for other modules
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)
# END OF IMPORT FIX

# Now, the import for data_pull should work as it's expected to be in the same directory
from data_pull import pull_data, find_column_match

import pandas as pd # Moved pandas import after the fix for consistency
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import DataBar, FormatObject, Rule


# Define the folder path for the output report
output_folder_path = r"C:\Reporting\Data Downloaded from IFS"

def perform_calculations():
    print("--- Starting perform_calculations() ---")
    ae_data, pt_data, p_data, project_manager_mapping = pull_data()

    pt_grouped = pd.DataFrame()
    if not pt_data.empty:
        activity_seq_col_pt = find_column_match(pt_data, 'Activity Seq')

        if not activity_seq_col_pt:
            print("Error: Could not find 'Activity Seq' (or similar) column in PT data. Cannot aggregate actual costs.")
        else:
            print(f"Using '{activity_seq_col_pt}' as the Activity Seq column for PT data aggregation.")

            cost_column_candidates = ['Total Internal Price', 'Internal Price', 'Sales Amount',
                                      'Sales Price', 'Internal Amount', 'Cost']
            actual_cost_col_pt = None
            for col_candidate in cost_column_candidates:
                match = find_column_match(pt_data, col_candidate)
                if match:
                    actual_cost_col_pt = match
                    break

            if actual_cost_col_pt:
                print(f"Using '{actual_cost_col_pt}' as the cost column for PT data aggregation.")
                pt_data[actual_cost_col_pt] = pd.to_numeric(pt_data[actual_cost_col_pt], errors='coerce')
                pt_data[actual_cost_col_pt].fillna(0, inplace=True)

                pt_grouped = pt_data.groupby(activity_seq_col_pt).agg({
                    actual_cost_col_pt: 'sum'
                }).reset_index()

                pt_grouped.rename(columns={activity_seq_col_pt: 'Activity Seq',
                                           actual_cost_col_pt: 'Actual Cost'}, inplace=True)
                print(f"Aggregated actual costs for {len(pt_grouped)} Activity Seq records from PT data.")
            else:
                print("Warning: No suitable cost column found in PT data. 'Actual Cost' will be missing or 0.")
    else:
        print("No PT data found to aggregate costs.")

    final_report = pd.DataFrame()
    if not ae_data.empty:
        final_report = ae_data.copy()
        if not pt_grouped.empty and 'Activity Seq' in final_report.columns and 'Activity Seq' in pt_grouped.columns:
            final_report = pd.merge(final_report, pt_grouped, on='Activity Seq', how='left')
            final_report['Actual Cost'] = final_report['Actual Cost'].fillna(0)
            print("Merged AE data with aggregated PT data.")
        else:
            final_report['Actual Cost'] = 0
            print("No PT data to merge or 'Activity Seq' missing for merge; 'Actual Cost' set to 0.")

        if 'Estimated Cost' in final_report.columns:
            final_report['Estimated Cost'] = pd.to_numeric(final_report['Estimated Cost'], errors='coerce').fillna(0)
        else:
            final_report['Estimated Cost'] = 0
            print("Warning: 'Estimated Cost' column not found in AE data. 'Budget Remaining' might be incorrect.")

        final_report['Budget Remaining'] = final_report['Estimated Cost'] - final_report['Actual Cost']
        print(f"Calculated 'Budget Remaining'. Final report has {len(final_report)} records before P data merge.")

    elif not pt_grouped.empty:
        print("Warning: AE data is empty. Report will be based on PT data only.")
        final_report = pt_grouped.copy()
        ae_cols_expected = ['Project', 'Project Description', 'Activity', 'Activity Description',
                            'Estimated Revenue', 'Estimated Cost']
        for col in ae_cols_expected:
            if col not in final_report.columns: final_report[col] = pd.NA
        if 'Estimated Cost' not in final_report.columns: final_report['Estimated Cost'] = 0
        if 'Actual Cost' not in final_report.columns: final_report['Actual Cost'] = 0
        final_report['Budget Remaining'] = final_report['Estimated Cost'] - final_report['Actual Cost']
    else:
        print("No AE or PT data available for the final report. Report will be empty.")
        final_report = pd.DataFrame()

    if not final_report.empty and project_manager_mapping and 'Project' in final_report.columns:
        final_report['Manager Description'] = final_report['Project'].astype(str).str.strip().map(project_manager_mapping).fillna('Unknown Manager')

        cols = list(final_report.columns)
        if 'Project Description' in cols and 'Manager Description' in cols:
            try:
                manager_desc_idx = cols.index('Manager Description')
                proj_desc_idx = cols.index('Project Description')
                if manager_desc_idx != proj_desc_idx +1: # Only move if not already in place
                    cols.pop(manager_desc_idx)
                    cols.insert(proj_desc_idx + 1, 'Manager Description')
                    final_report = final_report[cols]
                print("Added and reordered 'Manager Description' in the final report.")
            except ValueError:
                print("Could not reorder 'Manager Description' column as expected.")
        else:
            print("'Manager Description' added, but 'Project Description' not found for reordering.")

    elif not final_report.empty:
        final_report['Manager Description'] = 'Unknown Manager'
        print("Project manager mapping was empty or 'Project' column missing; 'Manager Description' set to 'Unknown Manager'.")

    employee_hours = pd.DataFrame()
    if not pt_data.empty:
        try:
            eh_cols_map = {
                'Internal Quantity': 'Internal Quantity',
                'Report Code Description': 'Report Code Description',
                'Project Activity Sequence': 'Activity Seq',
                'Employee Description': 'Employee Description'
            }
            eh_actual_cols = {}
            for hr_name, search_name in eh_cols_map.items():
                match = find_column_match(pt_data, search_name)
                if match:
                    eh_actual_cols[hr_name] = match
                else:
                    print(f"Warning: For Employee Hours, cannot find PT column for '{hr_name}' (searched for '{search_name}')")

            if eh_actual_cols:
                temp_eh_df = pd.DataFrame()
                for hr_name, actual_col in eh_actual_cols.items():
                    temp_eh_df[hr_name] = pt_data[actual_col]

                if not ae_data.empty and 'Activity Seq' in ae_data.columns and \
                   'Project Activity Sequence' in temp_eh_df.columns and \
                   'Project Description' in ae_data.columns:

                    ae_data_subset = ae_data[['Activity Seq', 'Project Description']].drop_duplicates(subset=['Activity Seq'])
                    
                    # Convert merge keys to string to avoid type mismatches if necessary
                    # temp_eh_df['Project Activity Sequence'] = temp_eh_df['Project Activity Sequence'].astype(str)
                    # ae_data_subset['Activity Seq'] = ae_data_subset['Activity Seq'].astype(str)

                    employee_hours = pd.merge(
                        temp_eh_df,
                        ae_data_subset,
                        left_on='Project Activity Sequence',
                        right_on='Activity Seq',
                        how='left'
                    )
                    if 'Activity Seq' in employee_hours.columns and 'Project Activity Sequence' in employee_hours.columns and \
                       employee_hours.columns.get_loc('Activity Seq') != employee_hours.columns.get_loc('Project Activity Sequence'): # if they are different columns
                         employee_hours.drop(columns=['Activity Seq'], inplace=True, errors='ignore')
                    print("Merged Employee Hours data with AE data for Project Description.")
                else:
                    employee_hours = temp_eh_df
                    employee_hours['Project Description'] = "N/A (AE link failed)"
                    print("Could not link Employee Hours to AE data for Project Description.")

                desired_eh_final_cols = ['Internal Quantity', 'Report Code Description', 'Project Description', 'Project Activity Sequence', 'Employee Description']
                for col in desired_eh_final_cols:
                    if col not in employee_hours.columns:
                        employee_hours[col] = pd.NA
                employee_hours = employee_hours[desired_eh_final_cols]
                print(f"Created Employee Hours DataFrame with {len(employee_hours)} records.")
            else:
                print("Not enough columns found in PT data to create Employee Hours report.")
        except Exception as e:
            print(f"Error creating Employee Hours DataFrame: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("No PT data available for Employee Hours report.")

    if not final_report.empty or not employee_hours.empty:
        output_file = os.path.join(output_folder_path, "reportX.xlsx")

        if not final_report.empty:
            cols = list(final_report.columns)
            if 'Estimated Revenue' in cols and 'Estimated Cost' in cols:
                try:
                    rev_idx = cols.index('Estimated Revenue')
                    cost_idx = cols.index('Estimated Cost')
                    if rev_idx != cost_idx:
                        cols[rev_idx], cols[cost_idx] = cols[cost_idx], cols[rev_idx]
                        final_report = final_report[cols]
                        print("Swapped 'Estimated Revenue' and 'Estimated Cost' columns.")
                except ValueError:
                     print("Could not swap 'Estimated Revenue' and 'Estimated Cost' columns.")

            sort_keys = []
            if 'Project' in final_report.columns: sort_keys.append('Project')
            if 'Budget Remaining' in final_report.columns: sort_keys.append('Budget Remaining')

            if sort_keys:
                final_report.sort_values(by=sort_keys, inplace=True, na_position='last')
                print(f"Sorted final report by {sort_keys}.")
            elif 'Activity Seq' in final_report.columns:
                 final_report.sort_values('Activity Seq', inplace=True, na_position='last')
                 print("Sorted final report by 'Activity Seq'.")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if not final_report.empty:
                final_report.to_excel(writer, sheet_name='Activity Report', index=False)
                print(f"Written 'Activity Report' sheet with {len(final_report)} rows.")

            used_sheet_names = {'Activity Report', 'Employee Hours'}
            managers_for_tabs = []
            if not final_report.empty and 'Manager Description' in final_report.columns:
                managers_for_tabs = [m for m in final_report['Manager Description'].unique() if pd.notna(m) and str(m).strip() not in ['', 'Unknown Manager']]

                if not managers_for_tabs:
                    print("No valid manager descriptions for separate tabs.")
                else:
                    print(f"Preparing to create separate tabs for {len(managers_for_tabs)} managers.")
                    for manager in managers_for_tabs:
                        manager_str = str(manager)
                        base_name = re.sub(r'[\\/*?:\[\]]', '_', manager_str)[:28].strip()
                        if not base_name: base_name = "UnnamedMgr"

                        sheet_name = base_name
                        counter = 1
                        while sheet_name in used_sheet_names:
                            sheet_name = f"{base_name[:26]}_{counter}"
                            counter += 1
                        used_sheet_names.add(sheet_name)

                        manager_data = final_report[final_report['Manager Description'] == manager_str]
                        if not manager_data.empty:
                            manager_data.to_excel(writer, sheet_name=sheet_name, index=False)
                            project_list = manager_data['Project'].unique().tolist() if 'Project' in manager_data else []
                            print(f"  - Created tab '{sheet_name}' for manager '{manager_str}' with {len(manager_data)} records. Projects: {project_list[:3]}...")
                        else:
                             print(f"  - Skipped tab for manager '{manager_str}' (no data after filtering).")

            if not employee_hours.empty:
                employee_hours.to_excel(writer, sheet_name='Employee Hours', index=False)
                print(f"Written 'Employee Hours' sheet with {len(employee_hours)} records.")

            workbook = writer.book
            def apply_formatting_to_worksheet(worksheet, data_df, apply_databar=True):
                if worksheet is None or data_df.empty:
                    return

                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                for i, col_name in enumerate(data_df.columns):
                    try:
                        col_letter = chr(65 + i)
                        if i >= 26: # Handle columns beyond Z
                            col_letter = chr(64 + (i // 26)) + chr(65 + (i % 26))
                        
                        max_len = data_df[col_name].astype(str).map(len).max()
                        header_len = len(str(col_name))
                        adjusted_width = max(max_len, header_len) + 2
                        worksheet.column_dimensions[col_letter].width = min(adjusted_width, 50)
                    except Exception as e:
                        print(f"Error adjusting width for column {col_name} ({col_letter}): {e}")


                currency_cols = ['Estimated Cost', 'Estimated Revenue', 'Actual Cost', 'Budget Remaining', 'Internal Quantity']
                for r_idx_plus_2, row_cells in enumerate(worksheet.iter_rows(min_row=2, max_row=len(data_df)+1), start=2):
                    for c_idx, cell in enumerate(row_cells):
                        if c_idx < len(data_df.columns): # Ensure column index is valid
                            col_name = data_df.columns[c_idx]
                            if col_name in currency_cols:
                                cell.number_format = '$#,##0.00'

                if apply_databar and 'Budget Remaining' in data_df.columns and not data_df.empty:
                    try:
                        br_col_idx = data_df.columns.get_loc('Budget Remaining')
                        br_col_letter = chr(65 + br_col_idx)
                        if br_col_idx >= 26:
                             br_col_letter = chr(64 + (br_col_idx // 26)) + chr(65 + (br_col_idx % 26))

                        data_bar_rule = Rule(type='dataBar',
                                             dataBar=DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')],
                                                             color="63C384"))
                        range_str = f"{br_col_letter}2:{br_col_letter}{len(data_df)+1}"
                        worksheet.conditional_formatting.add(range_str, data_bar_rule)
                    except Exception as e:
                        print(f"Error applying databar to {worksheet.title} col {br_col_letter}: {e}")


            if not final_report.empty and 'Activity Report' in workbook.sheetnames:
                apply_formatting_to_worksheet(workbook['Activity Report'], final_report, apply_databar=True)
                print("Applied formatting to 'Activity Report'.")

            if not employee_hours.empty and 'Employee Hours' in workbook.sheetnames:
                apply_formatting_to_worksheet(workbook['Employee Hours'], employee_hours, apply_databar=False)
                print("Applied formatting to 'Employee Hours'.")

            if managers_for_tabs: # Check if we intended to create manager tabs
                for sheet_name_in_wb in workbook.sheetnames:
                    if sheet_name_in_wb not in ['Activity Report', 'Employee Hours']: # This must be a manager sheet
                        # Find original manager name for this sheet to filter data correctly for formatting
                        original_manager_name_for_sheet = None
                        for manager_candidate in managers_for_tabs:
                            sanitized_candidate_base = re.sub(r'[\\/*?:\[\]]', '_', str(manager_candidate))[:28].strip()
                            if not sanitized_candidate_base: sanitized_candidate_base = "UnnamedMgr"
                            
                            # Check if sheet_name_in_wb matches base or base_counter
                            if sheet_name_in_wb == sanitized_candidate_base:
                                original_manager_name_for_sheet = manager_candidate
                                break
                            # Check for suffixed names (e.g., UnnamedMgr_1)
                            if "_" in sheet_name_in_wb:
                                prefix_sheet_name = sheet_name_in_wb.rsplit('_',1)[0]
                                if prefix_sheet_name == sanitized_candidate_base[:26]: # Match against potentially trimmed prefix
                                     original_manager_name_for_sheet = manager_candidate
                                     break


                        if original_manager_name_for_sheet:
                            manager_data_for_sheet = final_report[final_report['Manager Description'] == original_manager_name_for_sheet]
                            if not manager_data_for_sheet.empty:
                                apply_formatting_to_worksheet(workbook[sheet_name_in_wb], manager_data_for_sheet, apply_databar=True)
                                print(f"Applied formatting to manager sheet '{sheet_name_in_wb}'.")
                        else:
                             print(f"Could not reliably match sheet '{sheet_name_in_wb}' back to an original manager for formatting data.")
        print(f"Report successfully created: {output_file}")
    else:
        print("No data (neither final_report nor employee_hours) was available to write to the Excel report.")

if __name__ == "__main__":
    perform_calculations()