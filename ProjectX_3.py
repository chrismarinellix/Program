import os
import pandas as pd
import glob
import re
from openpyxl.styles import Font, PatternFill

# Define the folder path
folder_path = r"C:\py\Program\Data"

# Function to normalize column names for better matching
def normalize_column_name(name):
    return re.sub(r'\s+', '', name).lower()

# Function to find the best match for a column name
def find_column_match(df, column_name):
    # Normalize the target column name
    norm_target = normalize_column_name(column_name)
    
    # Create a dictionary of normalized column names to actual column names
    norm_to_actual = {normalize_column_name(col): col for col in df.columns}
    
    # Check for exact match after normalization
    if norm_target in norm_to_actual:
        return norm_to_actual[norm_target]
    
    # Check for partial matches
    possible_matches = [actual for norm, actual in norm_to_actual.items() 
                       if norm_target in norm or norm in norm_target]
    
    if possible_matches:
        return possible_matches[0]
    
    return None

# Function to read a file based on its extension
def read_file(file_path):
    _, ext = os.path.splitext(file_path)
    
    if ext.lower() in ['.csv']:
        try:
            return pd.read_csv(file_path)
        except UnicodeDecodeError:
            try:
                return pd.read_csv(file_path, encoding='latin1')
            except Exception as e:
                print(f"Error reading {file_path}: {e}")
                return pd.DataFrame()
    elif ext.lower() in ['.xlsx', '.xls']:
        try:
            return pd.read_excel(file_path)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return pd.DataFrame()
    elif ext.lower() in ['.txt', '.dat']:
        # Try with different separators
        for sep in ['\t', ',', ';', '|']:
            try:
                df = pd.read_csv(file_path, sep=sep)
                if len(df.columns) > 1:  # If more than one column, separator worked
                    return df
            except:
                continue
        print(f"Could not determine separator for {file_path}")
        return pd.DataFrame()
    else:
        print(f"Unsupported file format: {file_path}")
        return pd.DataFrame()

# Function to find AE, PT, and P files
def find_files(folder, file_type):
    all_files = []
    for extension in ['*.csv', '*.xlsx', '*.xls', '*.txt', '*.dat']:
        search_pattern = os.path.join(folder, f"*{file_type}*{extension}")
        all_files.extend(glob.glob(search_pattern))
    return all_files

# Find AE, PT, and P files
ae_files = find_files(folder_path, "AE")
pt_files = find_files(folder_path, "PT")
p_files = find_files(folder_path, "P")

print(f"Found {len(ae_files)} AE files: {ae_files}")
print(f"Found {len(pt_files)} PT files: {pt_files}")
print(f"Found {len(p_files)} P files: {p_files}")

# Read all AE files and concatenate them
ae_data_frames = []
for file in ae_files:
    print(f"Reading AE file: {file}")
    df = read_file(file)
    if not df.empty:
        print(f"  - Read {len(df)} rows with {len(df.columns)} columns")
        ae_data_frames.append(df)

# Concatenate all AE data frames
if ae_data_frames:
    ae_data = pd.concat(ae_data_frames, ignore_index=True)
    print(f"Total AE records: {len(ae_data)}")
else:
    ae_data = pd.DataFrame()
    print("No AE data loaded.")

# Read all PT files and concatenate them
pt_data_frames = []
for file in pt_files:
    print(f"Reading PT file: {file}")
    df = read_file(file)
    if not df.empty:
        print(f"  - Read {len(df)} rows with {len(df.columns)} columns")
        pt_data_frames.append(df)

# Concatenate all PT data frames
if pt_data_frames:
    pt_data = pd.concat(pt_data_frames, ignore_index=True)
    print(f"Total PT records: {len(pt_data)}")
else:
    pt_data = pd.DataFrame()
    print("No PT data loaded.")

# Read all P files and concatenate them
p_data_frames = []
for file in p_files:
    print(f"Reading P file: {file}")
    df = read_file(file)
    if not df.empty:
        print(f"  - Read {len(df)} rows with {len(df.columns)} columns")
        p_data_frames.append(df)

# Concatenate all P data frames
if p_data_frames:
    p_data = pd.concat(p_data_frames, ignore_index=True)
    print(f"Total P records: {len(p_data)}")
else:
    p_data = pd.DataFrame()
    print("No P data loaded.")

# Extract project manager information from P data
project_manager_mapping = {}
if not p_data.empty:
    try:
        # Find the relevant columns in P data
        project_col = find_column_match(p_data, 'Project')
        manager_desc_col = find_column_match(p_data, 'Manager Description')
        
        if project_col and manager_desc_col:
            print(f"Found project column '{project_col}' and manager description column '{manager_desc_col}'")
            
            # Create a mapping of project to manager description
            for _, row in p_data.iterrows():
                project = row[project_col]
                manager_desc = row[manager_desc_col]
                if pd.notna(project) and pd.notna(manager_desc):
                    project_manager_mapping[str(project).strip()] = str(manager_desc).strip()
            
            print(f"Created mapping for {len(project_manager_mapping)} projects to managers")
        else:
            print("Warning: Could not find Project or Manager Description columns in P data")
    except Exception as e:
        print(f"Error processing P data: {e}")
else:
    print("No P data available to extract manager information")

# Extract required fields from AE data
if not ae_data.empty:
    try:
        # Define required columns and look for matching columns
        ae_required_cols = {
            'Activity Seq': None,
            'Project': None,
            'Project Description': None,
            'Activity': None,
            'Activity Description': None,
            'Estimated Revenue': None,
            'Estimated Cost': None
        }
        
        # Find matching columns for each required column
        for req_col in ae_required_cols:
            match = find_column_match(ae_data, req_col)
            if match:
                ae_required_cols[req_col] = match
                print(f"Matched '{req_col}' to '{match}' in AE data")
            else:
                print(f"Warning: No match found for '{req_col}' in AE data")
        
        # Check if we found Activity Seq (primary key)
        if ae_required_cols['Activity Seq'] is None:
            print("Error: Could not find 'Activity Seq' column in AE data")
            ae_extract = pd.DataFrame()
        else:
            # Create a new DataFrame with matched columns
            ae_extract = pd.DataFrame()
            
            for req_col, actual_col in ae_required_cols.items():
                if actual_col:
                    ae_extract[req_col] = ae_data[actual_col]
                else:
                    # Create empty column if no match was found
                    ae_extract[req_col] = pd.NA
            
            # Group by Activity Seq to handle potential duplicates
            ae_extract = ae_extract.groupby('Activity Seq').first().reset_index()
            print(f"Extracted {len(ae_extract)} unique Activity Seq records from AE data")
    except Exception as e:
        print(f"Error processing AE data: {e}")
        ae_extract = pd.DataFrame()
else:
    print("No AE data found.")
    ae_extract = pd.DataFrame()

# Aggregate actual costs from PT data by Activity Seq
if not pt_data.empty:
    # Look for Activity Seq in PT data
    activity_seq_col = find_column_match(pt_data, 'Activity Seq')
    
    if not activity_seq_col:
        print("Error: Could not find 'Activity Seq' column in PT data")
        pt_grouped = pd.DataFrame()
    else:
        print(f"Using '{activity_seq_col}' as the Activity Seq column for PT data")
        
        # Try different possible column names for actual costs
        cost_columns = ['Total Internal Price', 'Internal Price', 'Sales Amount', 
                        'Sales Price', 'Internal Amount']
        
        cost_column = None
        for col in cost_columns:
            match = find_column_match(pt_data, col)
            if match:
                cost_column = match
                break
        
        if cost_column:
            print(f"Using '{cost_column}' as the cost column for PT data")
            
            # Aggregating by Activity Seq
            pt_grouped = pt_data.groupby(activity_seq_col).agg({
                cost_column: 'sum'
            }).reset_index()
            
            # Rename columns for consistency
            pt_grouped.rename(columns={activity_seq_col: 'Activity Seq', 
                                      cost_column: 'Actual Cost'}, inplace=True)
            
            print(f"Aggregated actual costs for {len(pt_grouped)} Activity Seq records from PT data")
        else:
            print(f"No suitable cost column found in PT data")
            pt_grouped = pd.DataFrame()
else:
    print("No PT data found.")
    pt_grouped = pd.DataFrame()

# Merge AE and PT data on Activity Seq
if not ae_extract.empty and not pt_grouped.empty:
    # Left join to keep all Activity Seq from AE data
    final_report = pd.merge(ae_extract, pt_grouped, on='Activity Seq', how='left')
    
    # Fill NaN values in Actual Cost with 0
    final_report['Actual Cost'] = final_report['Actual Cost'].fillna(0)
    
    # Calculate variance and rename to Budget Remaining
    final_report['Budget Remaining'] = final_report['Estimated Cost'] - final_report['Actual Cost']
    
    print(f"Created final report with {len(final_report)} records")
elif not ae_extract.empty:
    # If only AE data is available
    print("Only AE data available for the report")
    final_report = ae_extract
    final_report['Actual Cost'] = 0
    final_report['Budget Remaining'] = final_report['Estimated Cost']
elif not pt_grouped.empty:
    # If only PT data is available
    print("Only PT data available for the report")
    final_report = pt_grouped
    # Add missing columns
    for col in ['Project', 'Project Description', 'Activity', 'Activity Description', 
               'Estimated Revenue', 'Estimated Cost']:
        final_report[col] = pd.NA
    final_report['Budget Remaining'] = pd.NA
else:
    print("No data available for the report.")
    final_report = pd.DataFrame()

# Add Manager Description to the final report
if not final_report.empty and project_manager_mapping and 'Project' in final_report.columns:
    # Create a new Manager Description column
    final_report['Manager Description'] = final_report['Project'].apply(
        lambda x: project_manager_mapping.get(str(x).strip(), 'Unknown Manager') if pd.notna(x) else 'Unknown Manager'
    )
    
    # Move Manager Description column to be after Project Description
    cols = list(final_report.columns)
    if 'Project Description' in cols:
        proj_desc_idx = cols.index('Project Description')
        manager_desc_idx = cols.index('Manager Description')
        
        # Remove Manager Description from its current position
        cols.pop(manager_desc_idx)
        
        # Insert it after Project Description
        cols.insert(proj_desc_idx + 1, 'Manager Description')
        
        # Reorder the columns
        final_report = final_report[cols]
    
    print("Added Manager Description to the report")
else:
    print("Could not add Manager Description to the report")

# Write the final report to Excel if data is available
if not final_report.empty:
    output_file = os.path.join(folder_path, "reportX.xlsx")
    
    # Swap Estimated Revenue and Estimated Cost columns order
    cols = list(final_report.columns)
    if 'Estimated Revenue' in cols and 'Estimated Cost' in cols:
        rev_idx = cols.index('Estimated Revenue')
        cost_idx = cols.index('Estimated Cost')
        cols[rev_idx], cols[cost_idx] = cols[cost_idx], cols[rev_idx]
        final_report = final_report[cols]
    
    # Sort the report by Project and Budget Remaining (from smallest to largest)
    if 'Project' in final_report.columns and 'Budget Remaining' in final_report.columns:
        final_report.sort_values(['Project', 'Budget Remaining'], inplace=True)
    elif 'Activity Seq' in final_report.columns:
        final_report.sort_values('Activity Seq', inplace=True)
    
    # Create a Pandas Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the main report to the first sheet
        final_report.to_excel(writer, sheet_name='Activity Report', index=False)
        
        # Create separate tabs for each manager's projects
        if 'Manager Description' in final_report.columns:
            managers = final_report['Manager Description'].unique()
            print(f"Creating separate tabs for {len(managers)} managers")
            
            for manager in managers:
                if pd.isna(manager) or manager == 'Unknown Manager':
                    sheet_name = 'Unknown Manager'
                else:
                    # Clean the manager name for a valid sheet name (max 31 chars, no special chars)
                    sheet_name = str(manager)[:30].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                
                # Filter data for this manager
                manager_data = final_report[final_report['Manager Description'] == manager]
                
                if not manager_data.empty:
                    # Write to a new sheet
                    manager_data.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  - Added tab for manager: {manager} with {len(manager_data)} projects")
        
        # Apply formatting
        try:
            workbook = writer.book
            
            # Function to apply formatting to a worksheet
            def apply_formatting_to_worksheet(worksheet, data):
                # Style the header row
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust columns' width
                for i, col in enumerate(data.columns):
                    col_width = max(
                        data[col].astype(str).map(len).max(),
                        len(str(col))
                    ) + 2
                    column_letter = chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26)
                    worksheet.column_dimensions[column_letter].width = col_width
                
                # Format numeric columns
                for row in range(2, len(data) + 2):
                    for i, col in enumerate(data.columns):
                        column_letter = chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26)
                        cell = worksheet[f"{column_letter}{row}"]
                        
                        # Apply currency format to cost/revenue columns
                        if col in ['Estimated Cost', 'Estimated Revenue', 'Actual Cost', 'Budget Remaining']:
                            cell.number_format = '$#,##0.00'
                
                # Apply conditional formatting to Budget Remaining column if it exists
                if 'Budget Remaining' in data.columns and 'Project' in data.columns:
                    from openpyxl.formatting.rule import DataBar, FormatObject, Rule
                    
                    # Find the Budget Remaining column index
                    br_idx = list(data.columns).index('Budget Remaining')
                    br_col_letter = chr(65 + br_idx) if br_idx < 26 else chr(64 + br_idx//26) + chr(65 + br_idx%26)
                    
                    # Find the Project column index
                    project_idx = list(data.columns).index('Project')
                    
                    # Get unique projects with their start and end rows
                    unique_projects = []
                    current_project = None
                    start_row = 2  # Excel starts at row 1, row 2 is first data row after header
                    
                    # This worksheet data starts at index 1 (row 2 in Excel)
                    for i, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=project_idx+1, max_col=project_idx+1), start=2):
                        project_value = row[0].value
                        
                        if current_project is None:
                            current_project = project_value
                            start_row = i
                        elif project_value != current_project:
                            # End of current project group
                            unique_projects.append((current_project, start_row, i-1))
                            current_project = project_value
                            start_row = i
                    
                    # Add the last project group
                    if current_project is not None:
                        unique_projects.append((current_project, start_row, len(data)+1))
                    
                    # Apply conditional formatting for each project group
                    for proj_idx, (project, start_row, end_row) in enumerate(unique_projects):
                        # Create the green data bar conditional formatting
                        green_databar = DataBar(
                            cfvo=[FormatObject(type='min'), FormatObject(type='max')],
                            color="00B050",  # Green color
                            showValue=True,
                            minLength=0,
                            maxLength=100
                        )
                        
                        # Apply the conditional formatting to just this project's rows in the Budget Remaining column
                        rule = Rule(type='dataBar', dataBar=green_databar)
                        cell_range = f"{br_col_letter}{start_row}:{br_col_letter}{end_row}"
                        worksheet.conditional_formatting.add(cell_range, rule)
            
            # Apply formatting to main worksheet
            main_worksheet = writer.sheets['Activity Report']
            apply_formatting_to_worksheet(main_worksheet, final_report)
            
            # Apply formatting to each manager's worksheet
            if 'Manager Description' in final_report.columns:
                for manager in managers:
                    if pd.isna(manager) or manager == 'Unknown Manager':
                        sheet_name = 'Unknown Manager'
                    else:
                        sheet_name = str(manager)[:30].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                    
                    if sheet_name in writer.sheets:
                        manager_data = final_report[final_report['Manager Description'] == manager]
                        manager_worksheet = writer.sheets[sheet_name]
                        apply_formatting_to_worksheet(manager_worksheet, manager_data)
        
        except Exception as e:
            print(f"Warning: Could not apply formatting to Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"Report successfully created: {output_file}")
else:
    print("No data to write to the report.")

print("Process completed.")