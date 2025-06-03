"""
Generate an Activity/Cost report from AE, PT and P files using **Polars**.

**NEW in this version – always use the *latest* file per type**
----------------------------------------------------------------
* Only the most‑recent AE, PT and P file (by modified‑time) is processed.
* Temporary Excel lock files that start with "~$" are ignored.
* Empty / unreadable files are skipped so `pl.concat` never raises a width
  mismatch error.
* Updated for Polars ≥ 0.19 (renamed `.groupby()` → `.group_by()`).

Usage
-----
1. `pip install polars pandas openpyxl`
2. Edit `FOLDER_PATH` below (or pass it via `--folder`).
3. Run: `python ifs_report_polars.py`

The script creates **reportX.xlsx** in the same folder, with the main
"Activity Report" sheet plus one sheet per manager.
"""

from __future__ import annotations

import argparse
import os
import re
from pathlib import Path
from typing import Optional

import polars as pl
import pandas as pd
from openpyxl.formatting.rule import DataBar, FormatObject, Rule  # type: ignore
from openpyxl.styles import Font, PatternFill  # type: ignore

# ──────────────────────────────────────────────────────────────────────────────
# Optional shim: keep legacy code working on older Polars
# ──────────────────────────────────────────────────────────────────────────────
if not hasattr(pl.DataFrame, "groupby"):
    pl.DataFrame.groupby = pl.DataFrame.group_by  # type: ignore[attr-defined]

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def normalize(name: str) -> str:
    """Lower‑case, remove whitespace for fuzzy column matching."""
    return re.sub(r"\s+", "", str(name)).lower()


def find_col(df: pl.DataFrame, target: str) -> Optional[str]:
    """Return df column that best matches *target* (normalised)."""
    tgt = normalize(target)
    mapping = {normalize(c): c for c in df.columns}
    if tgt in mapping:
        return mapping[tgt]
    # partial matches
    for norm, actual in mapping.items():
        if tgt in norm or norm in tgt:
            return actual
    return None


def safe_read(path: Path) -> pl.DataFrame:
    """Best‑effort read using Polars; returns empty DF on failure."""
    ext = path.suffix.lower()
    try:
        if ext == ".csv":
            try:
                return pl.read_csv(path)
            except UnicodeDecodeError:
                return pl.read_csv(path, encoding="latin1")
        elif ext in {".xlsx", ".xls"}:
            try:
                return pl.read_excel(path)
            except Exception:
                return pl.from_pandas(pd.read_excel(path))
        elif ext in {".txt", ".dat"}:
            for sep in ["\t", ",", ";", "|"]:
                try:
                    df = pl.read_csv(path, separator=sep)
                    if df.width > 1:
                        return df
                except Exception:
                    pass
            print(f"⚠️  Could not detect separator for {path.name}")
            return pl.DataFrame()
        else:
            print(f"⚠️  Unsupported file {path.name}")
            return pl.DataFrame()
    except Exception as exc:
        print(f"⚠️  Failed reading {path.name}: {exc}")
        return pl.DataFrame()


def latest_file(folder: Path, tag: str) -> Optional[Path]:
    """Return most‑recent file containing *tag* (AE / PT / P)."""
    files: list[Path] = []
    for pattern in ["*.csv", "*.xlsx", "*.xls", "*.txt", "*.dat"]:
        files.extend(folder.glob(f"*{tag}*{pattern}"))
    # drop temp Excel locks
    files = [p for p in files if not p.name.startswith("~$")]
    return max(files, key=lambda p: p.stat().st_mtime) if files else None


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main(folder: str):
    root = Path(folder).expanduser().resolve()
    if not root.exists():
        raise SystemExit(f"Folder not found: {root}")

    print("Using source folder:", root)

    ae_path = latest_file(root, "AE")
    pt_path = latest_file(root, "PT")
    p_path = latest_file(root, "P")

    for lbl, p in (("AE", ae_path), ("PT", pt_path), ("P", p_path)):
        print(f"  {lbl}:", p.name if p else "❌ none found")

    AE = safe_read(ae_path) if ae_path else pl.DataFrame()
    PT = safe_read(pt_path) if pt_path else pl.DataFrame()
    P = safe_read(p_path) if p_path else pl.DataFrame()

    # ── project → manager mapping ─────────────────
    project_manager: dict[str, str] = {}
    if not P.is_empty():
        proj_col = find_col(P, "Project")
        mgr_col = find_col(P, "Manager Description")
        if proj_col and mgr_col:
            project_manager = dict(zip(P[proj_col].cast(str), P[mgr_col].cast(str)))

    # ── AE extract ────────────────────────────────
    required = [
        "Activity Seq",
        "Project",
        "Project Description",
        "Activity",
        "Activity Description",
        "Estimated Revenue",
        "Estimated Cost",
    ]
    if AE.is_empty():
        raise SystemExit("No AE data – aborting.")

    cols_map = {req: find_col(AE, req) for req in required}
    AE_EX = pl.DataFrame({
        req: AE[cols_map[req]] if cols_map[req] else pl.lit(None)
        for req in required
    })

    AE_EX = (
        AE_EX.group_by("Activity Seq", maintain_order=True)
        .agg(pl.all().first())
    )

    # ── PT aggregate ──────────────────────────────
    cost_candidates = [
        "Total Internal Price",
        "Internal Price",
        "Sales Amount",
        "Sales Price",
        "Internal Amount",
    ]
    PT_AGG = pl.DataFrame()
    if not PT.is_empty():
        act_col = find_col(PT, "Activity Seq")
        cost_col = next((find_col(PT, c) for c in cost_candidates if find_col(PT, c)), None)
        if act_col and cost_col:
            PT_AGG = (
                PT.group_by(act_col)
                .agg(pl.col(cost_col).sum().alias("Actual Cost"))
                .rename({act_col: "Activity Seq"})
            )

    # ── merge & compute ───────────────────────────
    FINAL = AE_EX.join(PT_AGG, on="Activity Seq", how="left")
    FINAL = FINAL.with_columns([
        pl.col("Actual Cost").fill_null(0),
        (pl.col("Estimated Cost") - pl.col("Actual Cost")).alias("Budget Remaining"),
    ])
    if project_manager:
        mgr_df = pl.DataFrame({
            "Project": list(project_manager),
            "Manager Description": list(project_manager.values()),
        })
        FINAL = (
            FINAL.join(mgr_df, on="Project", how="left")
            .with_columns(pl.col("Manager Description").fill_null("Unknown Manager"))
        )

    # swap Estimated Revenue / Cost order
    cols = FINAL.columns
    if {"Estimated Revenue", "Estimated Cost"}.issubset(cols):
        cols.remove("Estimated Revenue")
        idx = cols.index("Estimated Cost")
        cols.insert(idx, "Estimated Revenue")
        FINAL = FINAL.select(cols)

    # sort
    sort_cols = [c for c in ("Project", "Budget Remaining") if c in FINAL.columns] or ["Activity Seq"]
    FINAL = FINAL.sort(sort_cols)

    if FINAL.is_empty():
        raise SystemExit("No data to write.")

    write_excel(FINAL.to_pandas(), root / "reportX.xlsx")


# ──────────────────────────────────────────────────────────────────────────────
# Excel output + formatting
# ──────────────────────────────────────────────────────────────────────────────

def write_excel(df: pd.DataFrame, out_path: Path):
    print("Writing:", out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="Activity Report", index=False)
        if "Manager Description" in df.columns:
            for mgr, grp in df.groupby("Manager Description"):
                name = (
                    "Unknown Manager"
                    if pd.isna(mgr) or mgr == "Unknown Manager"
                    else str(mgr)[:30].translate(str.maketrans("/\\?*[]:", "________"))
                )
                grp.to_excel(xl, sheet_name=name, index=False)
        style_workbook(xl)
    print("✅ Excel saved.")


def style_workbook(writer: pd.ExcelWriter):
    wb = writer.book

    def style(ws, data: pd.DataFrame):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        # column widths
        for col_idx, col_name in enumerate(data.columns, 1):
            max_len = max(data[col_name].astype(str).map(len).max(), len(col_name)) + 2
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len
        money_cols = {"Estimated Cost", "Estimated Revenue", "Actual Cost", "Budget Remaining"}
        for row in ws.iter_rows(min_row=2, max_row=len(data) + 1):
            for cell in row:
                if data.columns[cell.column - 1] in money_cols:
                    cell.number_format = "$#,##0.00"
        # green data‑bar per project group on Budget Remaining
        if {"Project", "Budget Remaining"}.issubset(data.columns):
            br_idx = data.columns.get_loc("Budget Remaining") + 1
            pr_idx = data.columns.get_loc("Project") + 1
            current, start = None, 2
            for r in range(2, len(data) + 2):
                proj = ws.cell(row=r, column=pr_idx).value
                if current is None:
                    current = proj
                elif proj != current:
                    _add_bar(ws, br_idx, start, r - 1)
                    current, start = proj, r
            _add_bar(ws, br_idx, start, len(data) + 1)

    def _add_bar(ws, col_idx: int, r0: int, r1: int):
        bar = DataBar(cfvo=[FormatObject(type="min"), FormatObject(type="max")], color="00B050", showValue=True)
        rule = Rule(type="dataBar", dataBar=bar)
        col = ws.cell(row=1, column=col_idx).column_letter
        ws.conditional_formatting.add(f"{col}{r0}:{col}{r1}", rule)

    for sheet_name, ws in writer.sheets.items():
        # pandas passes values generator, so rebuild DataFrame for dimensions
        data = pd.DataFrame(ws.values)
        data.columns = data.iloc[0]
        data = data[1:].reset_index(drop=True)
        style(ws, data)


# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--folder", default=r"C:/Reporting/Data Downloaded from IFS", help="Root folder containing IFS downloads")
    args = ap.parse_args()
    main(args.folder)
